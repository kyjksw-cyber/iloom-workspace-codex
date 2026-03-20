#!/usr/bin/env python3
"""
일룸 매출 통합 대시보드 - 데이터 파이프라인
PRD Section 2-3 기반 ETL: raw-data → 정제 → 통합 → Google Sheets

Usage:
    python3 pipeline.py --raw-dir <path> --audit       # PRD vs 실제 파일 대조 리포트
    python3 pipeline.py --raw-dir <path> --spreadsheet-id <id>
    python3 pipeline.py --raw-dir <path> --create "일룸 매출 통합 대시보드"
    python3 pipeline.py --raw-dir <path> --dry-run     # 시트 입력 없이 검증만
"""

import sys
import os
import json
import csv
import argparse
import re
from pathlib import Path
from datetime import datetime

import openpyxl

# Google Sheets 연동
sys.path.insert(0, str(Path.home() / '.claude/lib'))

# ─────────────────────────────────────────────
# 1. 기준 데이터 로드
# ─────────────────────────────────────────────

def load_product_master(raw_dir):
    """상품기준시트 로드 → 코드 매핑 딕셔너리 생성"""
    path = raw_dir / "상품기준시트.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["상품마스터"]

    products = {}  # 일룸코드 → {상품명, 시리즈, 카테고리, 소비자가}
    naver_map = {}  # NV코드 → 일룸코드
    ohouse_map = {}  # OH코드 → 일룸코드
    official_map = {}  # 공식몰코드 → 일룸코드

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        products[code] = {
            "product_name": row[1],
            "series": row[2],
            "category": row[3],
            "consumer_price": int(row[4]) if row[4] else 0,
        }
        if row[5]:
            naver_map[row[5]] = code
        if row[6]:
            ohouse_map[row[6]] = code
        if row[7]:
            official_map[row[7]] = code

    return products, naver_map, ohouse_map, official_map


def load_store_master(raw_dir):
    """매장마스터 로드"""
    path = raw_dir / "매장마스터.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["매장목록"]

    stores = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        stores[code] = {
            "store_name": row[1],
            "region": row[2],
            "store_type": row[3],
        }
    return stores


def load_targets(raw_dir):
    """목표실적.csv 로드 → unpivot"""
    path = raw_dir / "목표실적.csv"
    targets = []

    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            store_code = row["매장코드"]
            for col, val in row.items():
                if col == "매장코드":
                    continue
                # col = "2026-01" 등
                targets.append({
                    "store_code": store_code,
                    "month": col,
                    "target": int(val),
                })
    return targets


# ─────────────────────────────────────────────
# 2. 채널별 원본 데이터 정제
# ─────────────────────────────────────────────

def parse_erp(raw_dir, products):
    """ERP 수주현황 정제: Row 1-4 제거, Row 5 헤더, 합계행 제거, 취소 제외"""
    records = []
    erp_files = sorted(raw_dir.glob("erp-매출-*.xlsx"))

    for path in erp_files:
        wb = openpyxl.load_workbook(path)
        ws = wb["수주현황"]

        for row in ws.iter_rows(min_row=6, values_only=True):
            seq = row[0]
            # 합계행: 순번이 숫자가 아닌 행
            if not isinstance(seq, (int, float)):
                continue

            status = row[14]  # 배송상태
            if status == "취소":
                continue

            date_val = str(row[2])[:10]  # 수주일자
            month = date_val[:7]
            product_code = row[4]  # 품목코드
            product_info = products.get(product_code, {})

            records.append({
                "date": date_val,
                "month": month,
                "channel": "ERP",
                "product_code": product_code,
                "product_name": row[5] or product_info.get("product_name", ""),
                "series": row[6] or product_info.get("series", ""),
                "category": row[7] or product_info.get("category", ""),
                "qty": int(row[8]) if row[8] else 0,
                "sale_amount": int(row[13]) if row[13] else 0,  # 수주금액
                "supply_price": int(row[10]) if row[10] else 0,  # 공급가
                "commission": 0,
                "settled_amount": int(row[13]) if row[13] else 0,  # ERP: 정산 = 수주금액
                "store_code": row[3],  # 대리점코드
            })

    return records


def parse_naver(raw_dir, products, naver_map):
    """네이버 주문내역 정제: 8컬럼 추출, 날짜 정리, 취소 제외, 코드 매핑"""
    records = []
    naver_files = sorted(raw_dir.glob("네이버-주문내역-*.xlsx"))

    for path in naver_files:
        wb = openpyxl.load_workbook(path)
        ws = wb["주문조회"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            order_status = row[17]  # 주문상태
            if order_status and "취소" in str(order_status):
                continue

            # 주문일시 → 날짜만
            datetime_str = str(row[2])
            date_val = datetime_str[:10]
            month = date_val[:7]

            nv_code = row[4]  # 상품번호 (NV-xxx)
            iloom_code = naver_map.get(nv_code, nv_code)
            product_info = products.get(iloom_code, {})

            sale_amount = int(row[10]) if row[10] else 0  # 상품별 합계금액
            settled = int(row[23]) if row[23] else 0  # 정산예정금액
            commission = int(row[24]) if row[24] else 0  # 수수료

            records.append({
                "date": date_val,
                "month": month,
                "channel": "네이버",
                "product_code": iloom_code,
                "product_name": row[5] or product_info.get("product_name", ""),
                "series": product_info.get("series", ""),
                "category": product_info.get("category", ""),
                "qty": int(row[7]) if row[7] else 0,
                "sale_amount": sale_amount,
                "supply_price": product_info.get("consumer_price", 0),  # 추정원가
                "commission": commission,
                "settled_amount": settled,
                "store_code": None,
            })

    return records


def parse_text_amount(val):
    """텍스트 금액 → 숫자: "1,890,000" → 1890000"""
    if val is None:
        return 0
    s = str(val).replace(",", "").strip()
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_ohouse(raw_dir, products, ohouse_map):
    """오늘의집 주문현황 정제: 텍스트→숫자, 날짜 통일, 취소 제외, 코드 매핑"""
    records = []
    oh_files = sorted(raw_dir.glob("오늘의집-주문현황-*.xlsx"))

    for path in oh_files:
        wb = openpyxl.load_workbook(path)
        ws = wb["주문현황"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            status = row[15]  # 배송상태
            if status and "취소" in str(status):
                continue

            # 날짜: 2026.01.20 → 2026-01-20
            date_raw = str(row[1]).replace(".", "-")
            date_val = date_raw[:10]
            month = date_val[:7]

            oh_code = row[3]  # 셀러상품코드 (OH-xxx)
            iloom_code = ohouse_map.get(oh_code, oh_code)
            product_info = products.get(iloom_code, {})

            sale_amount = parse_text_amount(row[9])  # 결제금액
            settled = parse_text_amount(row[17])  # 정산금액

            # 수수료율: "15%" → 0.15
            fee_rate_raw = str(row[18]).replace("%", "").strip() if row[18] else "0"
            try:
                fee_rate = float(fee_rate_raw) / 100 if float(fee_rate_raw) > 1 else float(fee_rate_raw)
            except ValueError:
                fee_rate = 0

            commission = int(sale_amount * fee_rate)

            records.append({
                "date": date_val,
                "month": month,
                "channel": "오늘의집",
                "product_code": iloom_code,
                "product_name": row[4] or product_info.get("product_name", ""),
                "series": product_info.get("series", ""),
                "category": product_info.get("category", ""),
                "qty": int(row[6]) if row[6] else 0,
                "sale_amount": sale_amount,
                "supply_price": product_info.get("consumer_price", 0),
                "commission": commission,
                "settled_amount": settled,
                "store_code": None,
            })

    return records


def parse_official(raw_dir, products, official_map):
    """공식몰 주문 정제: 날짜 변환, 취소 제외"""
    records = []
    csv_files = sorted(raw_dir.glob("공식몰-주문-*.csv"))

    for path in csv_files:
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("status", "").lower() in ("cancelled", "취소"):
                    continue

                # 날짜: MM/DD/YYYY → YYYY-MM-DD
                date_raw = row["order_date"]
                parts = date_raw.split("/")
                if len(parts) == 3:
                    date_val = f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
                else:
                    date_val = date_raw[:10]
                month = date_val[:7]

                product_code = row["product_code"]
                iloom_code = official_map.get(product_code, product_code)
                product_info = products.get(iloom_code, {})

                sale_amount = int(row["total_price"]) if row.get("total_price") else 0
                unit_price = int(row["unit_price"]) if row.get("unit_price") else 0

                records.append({
                    "date": date_val,
                    "month": month,
                    "channel": "공식몰",
                    "product_code": iloom_code,
                    "product_name": row.get("product_name", "") or product_info.get("product_name", ""),
                    "series": product_info.get("series", ""),
                    "category": product_info.get("category", ""),
                    "qty": int(row["qty"]) if row.get("qty") else 0,
                    "sale_amount": sale_amount,
                    "supply_price": unit_price,  # 공식몰: unit_price = 원가
                    "commission": 0,
                    "settled_amount": sale_amount,  # 자사몰: 수수료 없음
                    "store_code": None,
                })

    return records


# ─────────────────────────────────────────────
# 3. 통합 & 채산 계산
# ─────────────────────────────────────────────

def calculate_profitability(records):
    """margin, margin_rate 계산"""
    for r in records:
        r["margin"] = r["settled_amount"] - r["supply_price"]
        if r["sale_amount"] > 0:
            r["margin_rate"] = round(r["margin"] / r["sale_amount"] * 100, 2)
        else:
            r["margin_rate"] = 0
    return records


def enrich_store_info(records, stores):
    """매장마스터 조인: store_name, region 추가"""
    for r in records:
        if r["store_code"] and r["store_code"] in stores:
            info = stores[r["store_code"]]
            r["store_name"] = info["store_name"]
            r["region"] = info["region"]
        else:
            r["store_name"] = r["channel"] if not r["store_code"] else r["store_code"]
            r["region"] = "온라인" if not r["store_code"] else ""
    return records


# ─────────────────────────────────────────────
# 4. 시트별 요약 생성
# ─────────────────────────────────────────────

UNIFIED_COLS = [
    "date", "month", "channel", "product_code", "product_name",
    "series", "category", "qty", "sale_amount", "supply_price",
    "commission", "settled_amount", "margin", "margin_rate",
    "store_code", "store_name", "region",
]


def build_channel_summary(records):
    """채널별요약: 월 x 채널별 매출, 마진, 건수"""
    agg = {}
    for r in records:
        key = (r["month"], r["channel"])
        if key not in agg:
            agg[key] = {"sale_amount": 0, "margin": 0, "count": 0, "settled_amount": 0}
        agg[key]["sale_amount"] += r["sale_amount"]
        agg[key]["margin"] += r["margin"]
        agg[key]["settled_amount"] += r["settled_amount"]
        agg[key]["count"] += 1

    rows = []
    for (month, channel), v in sorted(agg.items()):
        margin_rate = round(v["margin"] / v["sale_amount"] * 100, 2) if v["sale_amount"] > 0 else 0
        rows.append([month, channel, v["sale_amount"], v["settled_amount"], v["margin"], margin_rate, v["count"]])

    return [["month", "channel", "sale_amount", "settled_amount", "margin", "margin_rate", "count"]] + rows


def build_store_summary(records, targets, stores):
    """매장별요약: 월별 매장 실적 + 목표 + 달성률"""
    # 매장별 월별 매출 집계 (ERP만)
    agg = {}
    for r in records:
        if r["channel"] != "ERP" or not r["store_code"]:
            continue
        key = (r["month"], r["store_code"])
        agg[key] = agg.get(key, 0) + r["sale_amount"]

    rows = []
    for t in targets:
        sc = t["store_code"]
        month = t["month"]
        actual = agg.get((month, sc), 0)
        target = t["target"]
        rate = round(actual / target * 100, 1) if target > 0 else 0

        store_info = stores.get(sc, {})
        rows.append([
            month, sc,
            store_info.get("store_name", sc),
            store_info.get("region", ""),
            store_info.get("store_type", ""),
            actual, target, rate,
        ])

    header = ["month", "store_code", "store_name", "region", "store_type", "actual", "target", "achievement_rate"]
    return [header] + sorted(rows, key=lambda x: (x[0], -x[7]))


# ─────────────────────────────────────────────
# 5. Google Sheets 출력
# ─────────────────────────────────────────────

def write_to_sheets(spreadsheet_id, unified_rows, channel_summary, store_summary):
    """구글시트에 3개 시트 데이터 입력"""
    from google_auth import get_credentials
    from googleapiclient.discovery import build

    creds = get_credentials()
    service = build('sheets', 'v4', credentials=creds)

    # 기존 시트 목록 확인
    spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing_sheets = {s['properties']['title']: s['properties']['sheetId']
                       for s in spreadsheet.get('sheets', [])}

    target_sheets = ["통합매출", "채널별요약", "목표달성"]

    # 시트 생성 (없으면) + 기존 데이터 클리어
    requests = []
    for name in target_sheets:
        if name not in existing_sheets:
            requests.append({"addSheet": {"properties": {"title": name}}})

    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id, body={"requests": requests}
        ).execute()
        # 새로 생성된 시트 ID 반영
        spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        existing_sheets = {s['properties']['title']: s['properties']['sheetId']
                           for s in spreadsheet.get('sheets', [])}

    # 기존 데이터 클리어
    ranges_to_clear = [f"{name}!A:Z" for name in target_sheets if name in existing_sheets]
    if ranges_to_clear:
        service.spreadsheets().values().batchClear(
            spreadsheetId=spreadsheet_id,
            body={"ranges": ranges_to_clear}
        ).execute()

    # 데이터 입력 (batch update)
    header = [UNIFIED_COLS]
    unified_data = header + [[r.get(c, "") for c in UNIFIED_COLS] for r in unified_rows]

    batch_data = [
        {"range": "통합매출!A1", "values": unified_data},
        {"range": "채널별요약!A1", "values": channel_summary},
        {"range": "목표달성!A1", "values": store_summary},
    ]

    service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"valueInputOption": "USER_ENTERED", "data": batch_data}
    ).execute()

    print(f"  통합매출: {len(unified_data)-1}행")
    print(f"  채널별요약: {len(channel_summary)-1}행")
    print(f"  목표달성: {len(store_summary)-1}행")


def create_spreadsheet(title):
    """새 스프레드시트 생성"""
    from google_auth import get_credentials
    from googleapiclient.discovery import build

    creds = get_credentials()
    service = build('sheets', 'v4', credentials=creds)

    body = {"properties": {"title": title}}
    result = service.spreadsheets().create(body=body).execute()
    return result["spreadsheetId"]


# ─────────────────────────────────────────────
# 6. 대시보드 JSON 출력 (선택)
# ─────────────────────────────────────────────

def export_dashboard_json(unified_rows, channel_summary, store_summary, output_dir):
    """대시보드용 JSON 파일 생성 (public/data/)"""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 통합매출 JSON
    with open(output_dir / "sales.json", "w", encoding="utf-8") as f:
        json.dump(unified_rows, f, ensure_ascii=False, indent=2)

    # 채널별요약 JSON
    ch_header = channel_summary[0]
    ch_data = [dict(zip(ch_header, row)) for row in channel_summary[1:]]
    with open(output_dir / "channel-summary.json", "w", encoding="utf-8") as f:
        json.dump(ch_data, f, ensure_ascii=False, indent=2)

    # 목표달성 JSON
    st_header = store_summary[0]
    st_data = [dict(zip(st_header, row)) for row in store_summary[1:]]
    with open(output_dir / "store-targets.json", "w", encoding="utf-8") as f:
        json.dump(st_data, f, ensure_ascii=False, indent=2)

    print(f"  JSON 출력: {output_dir}")


# ─────────────────────────────────────────────
# 7. PRD vs 실제 파일 대조 (Audit)
# ─────────────────────────────────────────────

PRD_SPEC = {
    "erp": {
        "pattern": "erp-매출-{mm}.xlsx",
        "sheet_name": "수주현황",
        "expected_cols": 20,
        "expected_rows_range": (50, 70),
        "header_row": 5,
        "junk_rows": [1, 2, 3, 4],
        "required_headers": ["순번", "수주번호", "수주일자", "대리점코드", "품목코드", "품목명",
                             "시리즈", "카테고리", "수량", "단가", "공급가", "소비자가",
                             "할인율", "수주금액", "배송상태"],
        "last_row_marker": "합계",
        "cancel_col_idx": 14,
        "cancel_value": "취소",
        "date_format": "YYYY-MM-DD",
    },
    "naver": {
        "pattern": "네이버-주문내역-{mm}.xlsx",
        "sheet_name": "주문조회",
        "expected_cols": 30,
        "expected_rows_range": (70, 90),
        "header_row": 1,
        "required_headers": ["주문일시", "상품번호", "상품명", "수량",
                             "상품별 합계금액", "정산예정금액", "수수료", "주문상태"],
        "cancel_col_idx": 17,
        "cancel_value": "취소",
        "date_format": "YYYY-MM-DD HH:MM:SS",
        "code_prefix": "NV-",
    },
    "ohouse": {
        "pattern": "오늘의집-주문현황-{mm}.xlsx",
        "sheet_name": "주문현황",
        "expected_cols": 20,
        "expected_rows_range": (40, 50),
        "header_row": 1,
        "required_headers": ["주문일", "셀러상품코드", "노출상품명", "주문수량",
                             "결제금액", "정산금액", "수수료율", "배송상태"],
        "cancel_col_idx": 15,
        "cancel_value": "취소",
        "date_format": "YYYY.MM.DD",
        "code_prefix": "OH-",
        "text_amount_cols": ["결제금액", "정산금액"],
        "percent_cols": ["수수료율"],
    },
    "official": {
        "pattern": "공식몰-주문-{mm}.csv",
        "expected_cols": 15,
        "expected_rows_range": (30, 40),
        "required_headers": ["order_date", "product_code", "product_name", "qty",
                             "unit_price", "total_price", "discount", "status"],
        "cancel_value": "cancelled",
        "date_format": "MM/DD/YYYY",
    },
}

MONTHS = ["202601", "202602", "202603"]


def run_audit(raw_dir):
    """PRD vs 실제 파일 대조 리포트"""
    findings = []  # (severity, category, message)
    OK = "OK"
    WARN = "WARN"
    FAIL = "FAIL"

    actual_files = set(os.listdir(raw_dir))

    # ── A. 파일 존재 여부 ──
    findings.append((None, "FILES", ""))  # section header

    for month in MONTHS:
        month_label = f"{month[:4]}-{month[4:]}"
        for ch, spec in PRD_SPEC.items():
            fname = spec["pattern"].format(mm=month)
            if fname in actual_files:
                findings.append((OK, "파일", f"{fname}"))
            else:
                findings.append((FAIL, "파일 누락", f"{fname} -- {month_label} {ch} 데이터 없음"))

    # 기준 데이터
    for name in ["상품기준시트.xlsx", "매장마스터.xlsx", "목표실적.csv"]:
        if name in actual_files:
            findings.append((OK, "기준 파일", name))
        else:
            findings.append((FAIL, "기준 파일 누락", name))

    # PRD에 없는 파일
    all_expected = set()
    for month in MONTHS:
        for spec in PRD_SPEC.values():
            all_expected.add(spec["pattern"].format(mm=month))
    all_expected.update(["상품기준시트.xlsx", "매장마스터.xlsx", "목표실적.csv"])
    unexpected = actual_files - all_expected - {".DS_Store"}
    for f in sorted(unexpected):
        findings.append((WARN, "미등록 파일", f"{f} -- PRD에 정의되지 않은 파일"))

    # ── B. 구조 검증 (존재하는 파일만) ──
    findings.append((None, "STRUCTURE", ""))

    for month in MONTHS:
        for ch, spec in PRD_SPEC.items():
            fname = spec["pattern"].format(mm=month)
            fpath = raw_dir / fname
            if not fpath.exists():
                continue

            month_label = f"{month[:4]}-{month[4:]}"
            prefix = f"[{ch}/{month_label}]"

            if fname.endswith(".xlsx"):
                wb = openpyxl.load_workbook(fpath)

                # 시트명 확인
                expected_sheet = spec.get("sheet_name")
                if expected_sheet:
                    if expected_sheet in wb.sheetnames:
                        findings.append((OK, "시트명", f"{prefix} '{expected_sheet}'"))
                    else:
                        findings.append((FAIL, "시트명 불일치",
                                         f"{prefix} PRD='{expected_sheet}' / 실제={wb.sheetnames}"))
                        continue

                ws = wb[expected_sheet] if expected_sheet else wb.active

                # 행/열 수
                rows, cols = ws.max_row, ws.max_column
                r_min, r_max = spec["expected_rows_range"]
                if r_min <= rows <= r_max:
                    findings.append((OK, "행수", f"{prefix} {rows}행 (PRD: {r_min}~{r_max})"))
                else:
                    findings.append((WARN, "행수 범위 밖",
                                     f"{prefix} {rows}행 (PRD: {r_min}~{r_max})"))

                if cols == spec["expected_cols"]:
                    findings.append((OK, "열수", f"{prefix} {cols}열"))
                else:
                    findings.append((WARN, "열수 불일치",
                                     f"{prefix} 실제={cols}열 / PRD={spec['expected_cols']}열"))

                # 헤더 컬럼 확인
                hrow = spec["header_row"]
                actual_headers = [c.value for c in ws[hrow] if c.value]
                missing = [h for h in spec["required_headers"] if h not in actual_headers]
                if not missing:
                    findings.append((OK, "필수 컬럼", f"{prefix} {len(spec['required_headers'])}개 모두 존재"))
                else:
                    findings.append((FAIL, "컬럼 누락",
                                     f"{prefix} 누락: {missing}"))

                # ERP 전용: Row 1-4 정크 확인
                if ch == "erp":
                    r1 = ws[1][0].value
                    if r1 and "리포트" in str(r1):
                        findings.append((OK, "ERP 정크행", f"{prefix} Row1 제목행 확인"))
                    else:
                        findings.append((WARN, "ERP 구조 변경",
                                         f"{prefix} Row1 = '{r1}' (PRD: 제목행 기대)"))

                    # 합계행 확인
                    last_val = ws[ws.max_row][0].value
                    if last_val == "합계":
                        findings.append((OK, "ERP 합계행", f"{prefix} 마지막행 '합계' 확인"))
                    else:
                        findings.append((WARN, "ERP 합계행 없음",
                                         f"{prefix} 마지막행 = '{last_val}'"))

                # 취소건 수
                cancel_idx = spec.get("cancel_col_idx")
                cancel_val = spec.get("cancel_value", "")
                if cancel_idx is not None:
                    data_start = spec["header_row"] + 1
                    cancel_count = 0
                    for row in ws.iter_rows(min_row=data_start, values_only=True):
                        if cancel_idx < len(row) and row[cancel_idx] and cancel_val in str(row[cancel_idx]):
                            cancel_count += 1
                    findings.append(("INFO", "취소건",
                                     f"{prefix} {cancel_count}건 (정제 시 제외됨)"))

                # 오늘의집 전용: 금액 텍스트 확인
                if ch == "ohouse":
                    r2 = [c.value for c in ws[2]]
                    amt_val = r2[9]  # 결제금액
                    if isinstance(amt_val, str) and "," in amt_val:
                        findings.append((OK, "금액 형식",
                                         f"{prefix} 결제금액=텍스트(쉼표) 확인 → 숫자 변환 필요"))
                    else:
                        findings.append((WARN, "금액 형식 변경",
                                         f"{prefix} 결제금액 타입={type(amt_val).__name__} (PRD: 쉼표 텍스트 기대)"))

                    fee_val = r2[18]  # 수수료율
                    if isinstance(fee_val, str) and "%" in fee_val:
                        findings.append((OK, "수수료율 형식",
                                         f"{prefix} 수수료율='{fee_val}' → % 파싱 필요"))
                    else:
                        findings.append((WARN, "수수료율 형식 변경",
                                         f"{prefix} 수수료율='{fee_val}' (PRD: '%' 텍스트 기대)"))

                    # 날짜 형식
                    date_val = str(r2[1])
                    if "." in date_val:
                        findings.append((OK, "날짜 형식", f"{prefix} '{date_val}' (점 구분) → 하이픈 변환 필요"))
                    else:
                        findings.append((WARN, "날짜 형식 변경",
                                         f"{prefix} '{date_val}' (PRD: 점 구분 기대)"))

            elif fname.endswith(".csv"):
                with open(fpath, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    rows_data = list(reader)
                    headers = reader.fieldnames or []

                row_count = len(rows_data)
                r_min, r_max = spec["expected_rows_range"]
                if r_min <= row_count <= r_max:
                    findings.append((OK, "행수", f"{prefix} {row_count}행"))
                else:
                    findings.append((WARN, "행수 범위 밖",
                                     f"{prefix} {row_count}행 (PRD: {r_min}~{r_max})"))

                col_count = len(headers)
                if col_count == spec["expected_cols"]:
                    findings.append((OK, "열수", f"{prefix} {col_count}열"))
                else:
                    findings.append((WARN, "열수 불일치",
                                     f"{prefix} 실제={col_count}열 / PRD={spec['expected_cols']}열"))

                missing = [h for h in spec["required_headers"] if h not in headers]
                if not missing:
                    findings.append((OK, "필수 컬럼", f"{prefix} {len(spec['required_headers'])}개 모두 존재"))
                else:
                    findings.append((FAIL, "컬럼 누락", f"{prefix} 누락: {missing}"))

                # 취소건
                cancel_count = sum(1 for r in rows_data
                                   if r.get("status", "").lower() == spec.get("cancel_value", ""))
                findings.append(("INFO", "취소건", f"{prefix} {cancel_count}건"))

                # 날짜 형식
                if rows_data:
                    date_sample = rows_data[0].get("order_date", "")
                    if "/" in date_sample:
                        findings.append((OK, "날짜 형식",
                                         f"{prefix} '{date_sample}' (MM/DD/YYYY) → YYYY-MM-DD 변환 필요"))

                    # 주소 쉼표 확인
                    comma_addr = sum(1 for r in rows_data if "," in r.get("address", ""))
                    if comma_addr:
                        findings.append((WARN, "CSV 파싱 주의",
                                         f"{prefix} 주소에 쉼표 포함 {comma_addr}건 (CSV 파싱 이슈 가능)"))

    # ── C. 데이터 무결성 검증 ──
    findings.append((None, "DATA_INTEGRITY", ""))

    # ERP 수주금액 검증
    for month in MONTHS:
        fname = f"erp-매출-{month}.xlsx"
        fpath = raw_dir / fname
        if not fpath.exists():
            continue

        month_label = f"{month[:4]}-{month[4:]}"
        wb = openpyxl.load_workbook(fpath)
        ws = wb["수주현황"]

        mismatch = 0
        total_amount = 0
        total_expected = 0
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not isinstance(row[0], (int, float)):
                continue
            qty = int(row[8]) if row[8] else 0
            supply = int(row[10]) if row[10] else 0
            disc = float(row[12]) if row[12] else 0
            amount = int(row[13]) if row[13] else 0
            expected = int(supply * qty * (1 - disc))
            total_amount += amount
            total_expected += expected
            if abs(amount - expected) > 1:
                mismatch += 1

        if mismatch > 0:
            ratio = total_amount / total_expected if total_expected else 0
            findings.append((FAIL, "ERP 수주금액 이상",
                             f"[erp/{month_label}] {mismatch}건 불일치 -- "
                             f"수주금액 합계={total_amount:,} vs "
                             f"공급가*qty*(1-할인율) 합계={total_expected:,} "
                             f"({ratio:.1f}x 차이). "
                             f"수주금액 컬럼의 계산 로직 확인 필요"))
        else:
            findings.append((OK, "ERP 수주금액", f"[erp/{month_label}] 공급가 기반 검증 통과"))

    # 코드 매핑 검증
    products, naver_map, ohouse_map, official_map = load_product_master(raw_dir)

    for month in MONTHS:
        # 네이버 코드 매핑
        nv_file = raw_dir / f"네이버-주문내역-{month}.xlsx"
        if nv_file.exists():
            wb = openpyxl.load_workbook(nv_file)
            ws = wb["주문조회"]
            unmapped = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[4]
                if code and code not in naver_map:
                    unmapped.add(code)
            if unmapped:
                findings.append((WARN, "코드 매핑 누락",
                                 f"[네이버/{month[:4]}-{month[4:]}] 상품기준시트에 없는 코드: {unmapped}"))
            else:
                findings.append((OK, "코드 매핑",
                                 f"[네이버/{month[:4]}-{month[4:]}] 전체 매핑 완료"))

        # 오늘의집 코드 매핑
        oh_file = raw_dir / f"오늘의집-주문현황-{month}.xlsx"
        if oh_file.exists():
            wb = openpyxl.load_workbook(oh_file)
            ws = wb["주문현황"]
            unmapped = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[3]
                if code and code not in ohouse_map:
                    unmapped.add(code)
            if unmapped:
                findings.append((WARN, "코드 매핑 누락",
                                 f"[오늘의집/{month[:4]}-{month[4:]}] 상품기준시트에 없는 코드: {unmapped}"))
            else:
                findings.append((OK, "코드 매핑",
                                 f"[오늘의집/{month[:4]}-{month[4:]}] 전체 매핑 완료"))

    # 매장코드 매핑
    stores = load_store_master(raw_dir)
    for month in MONTHS:
        erp_file = raw_dir / f"erp-매출-{month}.xlsx"
        if not erp_file.exists():
            continue
        wb = openpyxl.load_workbook(erp_file)
        ws = wb["수주현황"]
        unmapped_stores = set()
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not isinstance(row[0], (int, float)):
                continue
            sc = row[3]
            if sc and sc not in stores:
                unmapped_stores.add(sc)
        if unmapped_stores:
            findings.append((WARN, "매장코드 누락",
                             f"[erp/{month[:4]}-{month[4:]}] 매장마스터에 없는 코드: {unmapped_stores}"))
        else:
            findings.append((OK, "매장코드 매핑",
                             f"[erp/{month[:4]}-{month[4:]}] 전체 매핑 완료"))

    # 목표실적 매장코드 검증
    targets = load_targets(raw_dir)
    target_codes = set(t["store_code"] for t in targets)
    store_codes = set(stores.keys())
    missing_in_master = target_codes - store_codes
    missing_in_target = store_codes - target_codes
    if missing_in_master:
        findings.append((WARN, "목표실적 코드",
                         f"매장마스터에 없는 목표 매장코드: {missing_in_master}"))
    if missing_in_target:
        findings.append((WARN, "목표 미설정 매장",
                         f"목표실적에 없는 매장코드: {missing_in_target}"))
    if not missing_in_master and not missing_in_target:
        findings.append((OK, "목표 매장코드", f"매장마스터와 목표실적 코드 완전 일치 ({len(target_codes)}개)"))

    # ── D. 월별 커버리지 요약 ──
    findings.append((None, "COVERAGE", ""))
    for month in MONTHS:
        month_label = f"{month[:4]}-{month[4:]}"
        ch_status = []
        for ch, spec in PRD_SPEC.items():
            fname = spec["pattern"].format(mm=month)
            ch_status.append(f"{ch}={'O' if fname in actual_files else 'X'}")
        all_present = all("=O" in s for s in ch_status)
        sev = OK if all_present else FAIL
        findings.append((sev, f"{month_label}", " | ".join(ch_status)))

    # ── 리포트 출력 ──
    print("\n" + "=" * 72)
    print("  PRD vs 실제 파일 대조 리포트")
    print("=" * 72)

    current_section = None
    section_titles = {
        "FILES": "A. 파일 존재 여부",
        "STRUCTURE": "B. 구조 검증",
        "DATA_INTEGRITY": "C. 데이터 무결성",
        "COVERAGE": "D. 월별 커버리지",
    }

    ok_count = warn_count = fail_count = info_count = 0

    for sev, cat, msg in findings:
        if sev is None:
            print(f"\n--- {section_titles.get(cat, cat)} ---")
            current_section = cat
            continue

        icon = {"OK": "  [OK]", "WARN": "  [!!]", "FAIL": "  [XX]", "INFO": "  [--]"}.get(sev, "  ")
        print(f"{icon} {cat}: {msg}")

        if sev == OK: ok_count += 1
        elif sev == WARN: warn_count += 1
        elif sev == FAIL: fail_count += 1
        elif sev == "INFO": info_count += 1

    # 요약
    print(f"\n{'=' * 72}")
    print(f"  결과: OK={ok_count} | WARN={warn_count} | FAIL={fail_count} | INFO={info_count}")

    if fail_count > 0:
        print("  --> FAIL 항목 해결 후 파이프라인 실행 권장")
    elif warn_count > 0:
        print("  --> WARN 항목 확인 후 진행 가능 (데이터 품질 영향 가능)")
    else:
        print("  --> 모든 검증 통과. 파이프라인 실행 가능")

    print("=" * 72 + "\n")

    return fail_count


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="일룸 매출 통합 데이터 파이프라인")
    parser.add_argument("--raw-dir", required=True, help="raw-data 폴더 경로")
    parser.add_argument("--spreadsheet-id", help="기존 Google Sheets ID")
    parser.add_argument("--create", help="새 스프레드시트 생성 (제목)")
    parser.add_argument("--audit", action="store_true", help="PRD vs 실제 파일 대조 리포트")
    parser.add_argument("--dry-run", action="store_true", help="시트 입력 없이 검증만")
    parser.add_argument("--json-dir", help="대시보드 JSON 출력 경로")
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir)
    if not raw_dir.exists():
        print(f"Error: {raw_dir} 경로가 존재하지 않습니다.")
        sys.exit(1)

    # ── Audit 모드 ──
    if args.audit:
        fail_count = run_audit(raw_dir)
        sys.exit(1 if fail_count > 0 else 0)

    # ── Step 1: 기준 데이터 ──
    print("[1/6] 기준 데이터 로드...")
    products, naver_map, ohouse_map, official_map = load_product_master(raw_dir)
    stores = load_store_master(raw_dir)
    targets = load_targets(raw_dir)
    print(f"  상품: {len(products)}개, 매장: {len(stores)}개, 목표: {len(targets)}건")

    # ── Step 2: 채널별 정제 ──
    print("[2/6] 채널별 원본 데이터 정제...")
    erp = parse_erp(raw_dir, products)
    naver = parse_naver(raw_dir, products, naver_map)
    ohouse = parse_ohouse(raw_dir, products, ohouse_map)
    official = parse_official(raw_dir, products, official_map)
    print(f"  ERP: {len(erp)}건, 네이버: {len(naver)}건, 오늘의집: {len(ohouse)}건, 공식몰: {len(official)}건")

    # ── Step 3: 통합 ──
    print("[3/6] 채널 통합...")
    all_records = erp + naver + ohouse + official
    print(f"  통합: {len(all_records)}건")

    # ── Step 4: 채산 계산 ──
    print("[4/6] 채산(수익성) 계산...")
    all_records = calculate_profitability(all_records)
    all_records = enrich_store_info(all_records, stores)

    # ── Step 5: 요약 생성 ──
    print("[5/6] 시트별 요약 생성...")
    channel_summary = build_channel_summary(all_records)
    store_summary = build_store_summary(all_records, targets, stores)

    # 검증 출력
    months = sorted(set(r["month"] for r in all_records))
    channels = sorted(set(r["channel"] for r in all_records))
    print(f"  기간: {months}")
    print(f"  채널: {channels}")
    total_sales = sum(r["sale_amount"] for r in all_records)
    total_margin = sum(r["margin"] for r in all_records)
    print(f"  총 매출: {total_sales:,}원")
    print(f"  총 마진: {total_margin:,}원 ({total_margin/total_sales*100:.1f}%)" if total_sales > 0 else "")

    # ── Step 6: 출력 ──
    if args.dry_run:
        print("[6/6] Dry run -- 시트 입력 건너뜀")
    else:
        print("[6/6] Google Sheets 입력...")
        sid = args.spreadsheet_id

        if args.create:
            sid = create_spreadsheet(args.create)
            print(f"  새 스프레드시트 생성: {sid}")

        if not sid:
            print("Error: --spreadsheet-id 또는 --create 필요")
            sys.exit(1)

        write_to_sheets(sid, all_records, channel_summary, store_summary)
        print(f"  완료: https://docs.google.com/spreadsheets/d/{sid}")

    # JSON 출력 (선택)
    if args.json_dir:
        export_dashboard_json(all_records, channel_summary, store_summary, args.json_dir)

    print("\nDone.")


if __name__ == "__main__":
    main()
